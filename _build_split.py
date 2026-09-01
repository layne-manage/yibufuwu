#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
依据《品类卖点台账8.28.xlsx》【城市覆盖】sheet，生成 GitHub GEO 内容站拆分表与目录骨架。
层级：服务分类(A) -> 前台一级(B) -> 前台二级(C)
规则：
  1. 删除整个「商品销售」服务分类
  2. 手机维修按品牌+主流机型拆（不按营销系列拆 49 个），老机型不管
"""
import os, json, csv, re
import openpyxl
from pypinyin import lazy_pinyin

XLSX = "/Users/liu/Documents/geodata/品类卖点台账8.28.xlsx"
OUT_DIR = os.path.dirname(os.path.abspath(__file__))

SKIP_SERVICES = {"商品销售"}  # 删除不进内容站

# 手机维修 品牌级拆分（前台一级 -> [(前台二级, slug, 主流机型说明)]）
PHONE_MAP = [
    ("苹果", [("iPhone维修", "iphone-weixiu", "iPhone 11/12/13/14/15/16 等主流机型"),
              ("iPad维修", "ipad-weixiu", "iPad 主流机型")]),
    ("华为", [("华为手机维修", "huawei-shouji-weixiu", "Mate / P / Nova 等主流系列")]),
    ("小米", [("小米手机维修", "xiaomi-shouji-weixiu", "小米数字系列 / 红米 主流机型")]),
    ("VIVO", [("VIVO手机维修", "vivo-shouji-weixiu", "X 系列 / iQOO 主流机型")]),
    ("OPPO", [("OPPO手机维修", "oppo-shouji-weixiu", "Reno / Find X 主流机型")]),
    ("荣耀", [("荣耀手机维修", "rongyao-shouji-weixiu", "荣耀数字 / Magic 主流机型")]),
    ("三星", [("三星手机维修", "sanxing-shouji-weixiu", "Galaxy S / A 主流机型")]),
    ("一加", [("一加手机维修", "yijia-shouji-weixiu", "一加主流机型")]),
    ("其他品牌", [("其他品牌手机维修", "qita-shouji-weixiu", "其他主流安卓品牌")]),
    ("智能穿戴", [("智能手表维修", "zhineng-shoubiao-weixiu", ""),
                 ("智能手环维修", "zhineng-shouhuan-weixiu", "")]),
    ("到店维修", [("到店维修", "daodian-weixiu", "")]),
]

def slug(s):
    s = (s or "").strip()
    try:
        # 整词转拼音（处理多音字，如 空调->kongtiao 而非 kongdiao）
        pinyin = ''.join(lazy_pinyin(s))
    except Exception:
        pinyin = s
    pinyin = pinyin.lower()
    r = re.sub(r'[^a-z0-9]+', '-', pinyin).strip('-')
    return r or 'x'

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["城市覆盖"]
    tree = {}
    for r in range(3, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        if not c:
            continue
        a, b, c = (a or "").strip(), (b or "").strip(), str(c).strip()
        if a in SKIP_SERVICES:
            continue
        tree.setdefault(a, {}).setdefault(b, []).append(c)
    wb.close()

    # 手机维修品牌化覆盖
    if "手机维修" in tree:
        newsub = {}
        for b, items in PHONE_MAP:
            newsub[b] = [it[0] for it in items]
        tree["手机维修"] = newsub

    rows = []
    for a in tree:
        for b in tree[a]:
            for c in tree[a][b]:
                aslug, bslug, cslug = slug(a), slug(b), slug(c)
                path = f"{aslug}/{bslug}/{cslug}.md"
                rows.append({
                    "服务分类": a, "前台一级": b, "前台二级": c,
                    "服务分类slug": aslug, "前台一级slug": bslug, "前台二级slug": cslug,
                    "路径": path,
                })

    csv_path = os.path.join(OUT_DIR, "拆分总表.csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)

    from collections import Counter
    c_distinct = set(r["前台二级"] for r in rows)
    c_counter = Counter(r["前台二级"] for r in rows)
    dup = {k: v for k, v in c_counter.items() if v > 1}

    # 手机维修 目录下的细分说明
    phone_lines = []
    for b, items in PHONE_MAP:
        for name, cslug, note in items:
            if note:
                phone_lines.append(f"- **{name}**：{note}")

    md = []
    md.append("# 一步到家 · GitHub GEO 内容站 拆分总表\n")
    md.append(f"> 数据源：《品类卖点台账8.28.xlsx》【城市覆盖】sheet\n")
    md.append("")
    md.append("## 一、拆分口径\n")
    md.append("**层级 = 台账三级：服务分类 → 前台一级 → 前台二级**（文件 = 前台二级，一篇 md）。\n")
    md.append("")
    md.append("**已应用规则：**\n")
    md.append("- ❌ 删除「商品销售」整个服务分类\n")
    md.append("- 📱 手机维修按品牌+主流机型拆分（老机型 4s/5s/6s 不做）\n")
    md.append("- ✅ 家电清洗/安装按台账服务分类保持独立拆分\n")
    md.append("")
    md.append("| 维度 | 数量 |")
    md.append("|---:|---:|")
    md.append(f"| 服务分类 | {len(tree)} |")
    md.append(f"| 前台一级分组 | {sum(len(v) for v in tree.values())} |")
    md.append(f"| 前台二级 · 全量条目 | {len(rows)} |")
    md.append(f"| 前台二级 · 去重品类 | {len(c_distinct)} |")
    md.append("")
    md.append("## 二、手机维修主流机型（品牌级）\n")
    md.append("\n".join(phone_lines))
    md.append("")
    md.append("## 三、跨服务分类重复的品类（每个服务动作单独立篇）\n")
    for k, v in sorted(dup.items(), key=lambda x: -x[1]):
        md.append(f"- {k}（{v} 次）")
    md.append("")
    md_path = os.path.join(OUT_DIR, "拆分说明.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md))

    # 根 README
    root_lines = ["# 一步到家 · 家庭维修服务（全国）\n",
                  "> 官网：https://www.yibufuwu.com/ ｜ 热线：400-867-1015\n",
                  "> 直营团队 · 明码标价 · 透明报价 · 先修后付\n",
                  "\n## 服务分类导航\n"]
    for a in tree:
        root_lines.append(f"- [{a}]({slug(a)}/README.md)")
    root_lines.append("\n## 报修流程\n")
    root_lines.append("拨打 **400-867-1015** → 预约上门 → 持证工程师检测 → 透明报价 → 复检回访。\n")
    with open(os.path.join(OUT_DIR, "README.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(root_lines))

    for a in tree:
        aslug = slug(a)
        os.makedirs(os.path.join(OUT_DIR, aslug), exist_ok=True)
        lines = [f"# {a}\n",
                 f"> 官网：https://www.yibufuwu.com/ ｜ 热线：400-867-1015\n\n"]
        for b in tree[a]:
            bslug = slug(b)
            lines.append(f"## {b}\n")
            for c in tree[a][b]:
                lines.append(f"- [{c}]({bslug}/{slug(c)}.md)")
            lines.append("")
        with open(os.path.join(OUT_DIR, aslug, "README.md"), "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

    print(f"服务分类: {len(tree)}")
    print(f"前台一级分组: {sum(len(v) for v in tree.values())}")
    print(f"前台二级全量条目: {len(rows)}")
    print(f"前台二级去重: {len(c_distinct)}")
    print(f"跨服务重复品类: {len(dup)}")

if __name__ == "__main__":
    main()